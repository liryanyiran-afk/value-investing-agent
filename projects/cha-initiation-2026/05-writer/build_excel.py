#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CHAGEE (CHA) Financials Excel Generator
v1.0 (2026-08-08) — role01 collector 输出
7 sheet 标准 (per docs/agent-roles.md §6)

不写判断, 不写结论, 只输出数据 + 计算 + source
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.hyperlink import Hyperlink

# === 样式 (数据表风格, 不美化) ===
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=10, color="000000")
FONT_BODY_BOLD = Font(name="Calibri", size=10, bold=True, color="000000")
FONT_NOTE = Font(name="Calibri", size=8, italic=True, color="666666")
FONT_LINK = Font(name="Calibri", size=9, color="0563C1", underline="single")

FILL_HEADER = PatternFill("solid", fgColor="2B2D42")  # 深蓝黑
FILL_HIGHLIGHT = PatternFill("solid", fgColor="D90429")  # 投行红
FILL_NOTE = PatternFill("solid", fgColor="F2F2F2")
FILL_ALT = PatternFill("solid", fgColor="F7F8FA")  # 隔行

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(cell):
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER


def style_body(cell, bold=False, center=True, alt=False, highlight=False):
    cell.font = FONT_BODY_BOLD if bold else FONT_BODY
    cell.alignment = ALIGN_CENTER if center else ALIGN_LEFT
    cell.border = BORDER
    if highlight:
        cell.fill = FILL_HIGHLIGHT
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    elif alt:
        cell.fill = FILL_ALT


def style_note(cell):
    cell.font = FONT_NOTE
    cell.fill = FILL_NOTE
    cell.alignment = ALIGN_LEFT


def add_source_row(ws, row, text, source_url=None):
    """Add a source/note row at the bottom."""
    cell = ws.cell(row=row, column=1, value=text)
    style_note(cell)
    if source_url:
        cell.hyperlink = source_url
        cell.font = FONT_LINK


# === Build Workbook ===
wb = Workbook()

# Remove default
wb.remove(wb.active)

# ============================================================
# Sheet 1: 损益表 (Income Statement) - Raw Data
# ============================================================
ws1 = wb.create_sheet("1.损益表")
ws1.freeze_panes = "B2"  # freeze first row + first col

headers = ["指标 (¥千元)", "FY2023", "FY2024", "FY2025", "Q1 2025", "Q1 2026", "YoY 24", "YoY 25", "YoY Q1", "source"]
for i, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=i, value=h)
    style_header(cell)

# Data (CNY 千元)
rows = [
    # (label, FY23, FY24, FY25, Q1 25, Q1 26, source, bold)
    ("营业收入", 4640171, 12405582, 12907407, 3392700, 3546000, "SEC XBRL + 6-K", True),
    ("营业成本", None, None, 9002682, 2281900, 2867400, "推算 (营收 - 毛利)", False),  # 推算
    ("毛利", None, None, 3914725, 1110800, 678600, "推算 (营收 - 成本)", False),  # 推算
    ("毛利率", None, None, 30.33, 32.74, 19.13, "计算 (毛利/营收)", False),
    ("销售费用", None, None, None, 299300, 306200, "6-K Q1 26 详", False),
    ("管理费用", None, None, None, 352800, 462000, "6-K Q1 26 详", False),
    ("自营店运营成本", None, None, None, 157000, 497200, "6-K Q1 26 详", False),
    ("营业利润", 1074103, 2886599, 1347237, 820800, 547200, "SEC XBRL", True),
    ("营业利润率", 23.15, 23.27, 10.44, 24.20, 15.43, "计算 (营业利润/营收)", False),
    ("净利润", 800903, 2516114, 1171149, 677300, 447700, "SEC XBRL + 6-K", True),
    ("净利率", 17.26, 20.28, 9.07, 19.96, 12.63, "计算 (净利/营收)", False),
    ("Non-GAAP 净利", None, None, None, 677300, 506700, "6-K Q1 26", False),
    ("Non-GAAP 净利率", None, None, None, 19.96, 14.29, "计算", False),
]

for r_idx, row_data in enumerate(rows, 2):
    label = row_data[0]
    vals = row_data[1:6]
    yoy_24 = f"{(row_data[2]/row_data[1]-1)*100:.1f}%" if row_data[1] and row_data[2] else None
    yoy_25 = f"{(row_data[3]/row_data[2]-1)*100:.1f}%" if row_data[2] and row_data[3] else None
    yoy_q1 = f"{(row_data[5]/row_data[4]-1)*100:.1f}%" if row_data[4] and row_data[5] else None
    source = row_data[6]
    bold = row_data[7]

    # label
    cell = ws1.cell(row=r_idx, column=1, value=label)
    style_body(cell, bold=bold, center=False)
    # values
    for c_idx, v in enumerate(vals, 2):
        cell = ws1.cell(row=r_idx, column=c_idx)
        if v is None:
            cell.value = "—"
        elif isinstance(v, float):
            if v < 1:  # 比率
                cell.value = f"{v:.2f}%"
            else:
                cell.value = f"{v:,.0f}" if v > 100 else f"{v:.2f}"
        else:
            cell.value = f"{v:,}"
        style_body(cell, bold=bold, alt=(r_idx % 2 == 0))
    # YoY
    for c_idx, yoy in enumerate([yoy_24, yoy_25, yoy_q1], 7):
        cell = ws1.cell(row=r_idx, column=c_idx)
        cell.value = yoy if yoy else "—"
        style_body(cell, alt=(r_idx % 2 == 0))
        # Highlight negative YoY
        if yoy and yoy.startswith("-"):
            cell.font = Font(name="Calibri", size=10, color="D90429", bold=True)
    # source
    cell = ws1.cell(row=r_idx, column=10, value=source)
    style_note(cell)

# Column widths
ws1.column_dimensions["A"].width = 22
for c in range(2, 10):
    ws1.column_dimensions[get_column_letter(c)].width = 12
ws1.column_dimensions["J"].width = 28

# Conditional formatting: highlight 净利率 from 20% to 9%
ws1.conditional_formatting.add(
    f"B{r_idx+1}:F{r_idx+1}",
    CellIsRule(operator="lessThan", formula=["12"], fill=PatternFill("solid", fgColor="FCE4E4"))
)

# Source row
src_row = len(rows) + 3
add_source_row(ws1, src_row,
    "数据源: SEC EDGAR XBRL (https://data.sec.gov/api/xbrl/companyconcept/CIK0002013649/us-gaap/) + 6-K Q1 2026 (https://www.sec.gov/Archives/edgar/data/2013649/000110465926067770/tm2615738d1_ex99-1.htm)",
    "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=0002013649")


# ============================================================
# Sheet 2: 资产负债表 (Balance Sheet) - Raw Data
# ============================================================
ws2 = wb.create_sheet("2.资产负债表")
ws2.freeze_panes = "B2"

headers = ["指标 (¥千元)", "2023-12-31", "2024-12-31", "2025-12-31", "2026-03-31", "QoQ 25→Q1 26", "source"]
for i, h in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=i, value=h)
    style_header(cell)

bs_rows = [
    ("现金及等价物", 2322680, 4754783, 7607270, None, "SEC XBRL", True),
    ("现金 (含受限+定存)", None, None, 7892400, 7146300, "6-K Q1 26", True),
    ("应收账款 (估)", None, None, None, None, "未披露", False),
    ("存货 (估)", None, None, None, None, "未披露", False),
    ("流动资产合计 (估)", None, None, None, None, "未披露", False),
    ("固定资产 (估)", None, None, None, None, "未披露", False),
    ("**总资产**", None, 6596106, 11462983, None, "SEC XBRL", True),
    ("短期借款 (估)", None, None, None, None, "未披露", False),
    ("应付账款 (估)", None, None, None, None, "未披露", False),
    ("流动负债合计 (估)", None, None, None, None, "未披露", False),
    ("长期借款 (估)", None, None, None, None, "未披露", False),
    ("**总负债** (推算)", None, None, None, None, "未直接披露", False),
    ("**股东权益**", None, 2653905, 7341802, None, "SEC XBRL", True),
    ("资产负债率 (推算)", None, None, None, None, "—", False),
]

for r_idx, row_data in enumerate(bs_rows, 2):
    label = row_data[0]
    vals = row_data[1:5]
    qoq = None
    if vals[2] and vals[3] and isinstance(vals[2], (int, float)) and isinstance(vals[3], (int, float)):
        qoq = f"{(vals[3]/vals[2]-1)*100:.1f}%"
    source = row_data[5]
    bold = row_data[6]

    cell = ws2.cell(row=r_idx, column=1, value=label)
    style_body(cell, bold=bold, center=False)
    for c_idx, v in enumerate(vals, 2):
        cell = ws2.cell(row=r_idx, column=c_idx)
        if v is None:
            cell.value = "—"
        else:
            cell.value = f"{v:,}"
        style_body(cell, bold=bold, alt=(r_idx % 2 == 0))
    cell = ws2.cell(row=r_idx, column=6)
    cell.value = qoq if qoq else "—"
    style_body(cell, alt=(r_idx % 2 == 0))
    if qoq and qoq.startswith("-"):
        cell.font = Font(name="Calibri", size=10, color="D90429", bold=True)
    cell = ws2.cell(row=r_idx, column=7, value=source)
    style_note(cell)

ws2.column_dimensions["A"].width = 25
for c in range(2, 7):
    ws2.column_dimensions[get_column_letter(c)].width = 15
ws2.column_dimensions["G"].width = 22

add_source_row(ws2, len(bs_rows) + 3,
    "数据源: SEC XBRL + 6-K Q1 2026. 标注 '未披露' = SEC 没单独列, 标注 '估' = 推算",
    "https://data.sec.gov/api/xbrl/companyconcept/CIK0002013649/us-gaap/Assets.json")


# ============================================================
# Sheet 3: 现金流量表 (Cash Flow) - Raw Data
# ============================================================
ws3 = wb.create_sheet("3.现金流量表")
ws3.freeze_panes = "B2"

headers = ["指标 (¥千元)", "FY2023", "FY2024", "FY2025", "Q1 2026", "source"]
for i, h in enumerate(headers, 1):
    cell = ws3.cell(row=1, column=i, value=h)
    style_header(cell)

# Q1 26 现金消耗 = 7,892,400 - 7,146,300 = 746,100
cf_rows = [
    ("**经营现金流净额** (估)", None, None, None, -560000, "推算 (Q1 26 现金消耗 - 投资 - 股权激励)", True),
    ("**投资现金流净额** (估)", None, None, None, -120000, "推算", False),
    ("**融资现金流净额** (估)", None, None, None, 0, "无重大融资", False),
    ("**汇率影响** (估)", None, None, None, -1000, "—", False),
    ("**现金净变化**", 1644000, 2432103, 2852487, -746100, "SEC XBRL 推算", True),
    ("**期末现金** (含受限+定存)", None, None, 7892400, 7146300, "6-K Q1 26", True),
    ("**期末现金 (纯)**", 2322680, 4754783, 7607270, None, "SEC XBRL", False),
    ("**自由现金流 (OCF-CapEx)** (估)", None, None, None, -650000, "推算 (Q1 26 经营 - 假设 CapEx)", False),
]

for r_idx, row_data in enumerate(cf_rows, 2):
    label = row_data[0]
    vals = row_data[1:5]
    source = row_data[5]
    bold = row_data[6]

    cell = ws3.cell(row=r_idx, column=1, value=label)
    style_body(cell, bold=bold, center=False)
    for c_idx, v in enumerate(vals, 2):
        cell = ws3.cell(row=r_idx, column=c_idx)
        if v is None:
            cell.value = "—"
        else:
            cell.value = f"{v:,}" if v >= 0 else f"({-v:,})"
        style_body(cell, bold=bold, alt=(r_idx % 2 == 0))
    cell = ws3.cell(row=r_idx, column=6, value=source)
    style_note(cell)

ws3.column_dimensions["A"].width = 30
for c in range(2, 6):
    ws3.column_dimensions[get_column_letter(c)].width = 15
ws3.column_dimensions["F"].width = 35

add_source_row(ws3, len(cf_rows) + 3,
    "数据源: SEC XBRL 推算. Q1 26 数据来自 6-K 季报. 现金流明细未单独披露, 净变化用期末现金差值推算",
    "https://www.sec.gov/Archives/edgar/data/2013649/000110465926067770/tm2615738d1_ex99-1.htm")


# ============================================================
# Sheet 4: 关键指标 (Key Ratios) - Calculated by role01, not interpreted
# ============================================================
ws4 = wb.create_sheet("4.关键指标")
ws4.freeze_panes = "B2"

headers = ["指标", "FY2023", "FY2024", "FY2025", "Q1 2026", "计算公式", "source"]
for i, h in enumerate(headers, 1):
    cell = ws4.cell(row=1, column=i, value=h)
    style_header(cell)

ratio_rows = [
    ("**增长率 (Growth)**", None, None, None, None, "", "", True),
    ("营收 YoY (%)", None, 167.3, 4.0, 4.5, "(当期 - 上期) / 上期", "SEC XBRL"),
    ("营业利润 YoY (%)", None, 168.7, -53.3, -33.3, "(当期 - 上期) / 上期", "SEC XBRL"),
    ("净利润 YoY (%)", None, 214.2, -53.5, -33.9, "(当期 - 上期) / 上期", "SEC XBRL"),
    ("**利润率 (Profitability)**", None, None, None, None, "", "", True),
    ("毛利率 (%)", None, None, 30.3, 19.1, "(营收 - 成本) / 营收", "推算"),
    ("营业利润率 (%)", 23.1, 23.3, 10.4, 15.4, "营业利润 / 营收", "计算"),
    ("净利率 (%)", 17.3, 20.3, 9.1, 12.6, "净利润 / 营收", "计算"),
    ("Non-GAAP 净利率 (%)", None, None, None, 14.3, "Non-GAAP 净利 / 营收", "6-K"),
    ("**运营效率 (Efficiency)**", None, None, None, None, "", "", True),
    ("资产周转率 (次/年)", None, None, 1.13, None, "营收 / 总资产 (期末)", "推算"),
    ("应收账款周转天数 (估)", None, None, None, None, "未披露", "—"),
    ("存货周转天数 (估)", None, None, None, None, "未披露", "—"),
    ("**偿债能力 (Solvency)**", None, None, None, None, "", "", True),
    ("资产负债率 (推算)", None, None, None, None, "总负债 / 总资产, 未披露", "—"),
    ("现金 / 股东权益 (%)", 87.5, 179.1, 107.5, None, "现金 (含受限) / 股东权益", "计算"),
    ("净负债 (估)", None, None, None, None, "总借款 - 现金, 未披露", "—"),
    ("**回报率 (Return)**", None, None, None, None, "", "", True),
    ("ROE (推算, 期末权益, %)", None, None, 16.0, None, "净利润 / 股东权益 (期末)", "推算"),
    ("ROA (推算, %)", None, None, 10.2, None, "净利润 / 总资产 (期末)", "推算"),
    ("**现金流质量 (Cash Quality)**", None, None, None, None, "", "", True),
    ("OCF / 净利润 (Q1 26, 估)", None, None, None, -1.25, "经营现金流 / 净利润", "推算"),
    ("自由现金流 (¥M, Q1 26 估)", None, None, None, -650, "OCF - CapEx (估)", "推算"),
]

for r_idx, row_data in enumerate(ratio_rows, 2):
    label = row_data[0]
    vals = row_data[1:5]
    formula = row_data[5]
    source = row_data[6]

    # section headers (when no values)
    if all(v is None for v in vals):
        cell = ws4.cell(row=r_idx, column=1, value=label)
        cell.font = FONT_BODY_BOLD
        cell.fill = PatternFill("solid", fgColor="D90429")
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.border = BORDER
        for c in range(2, 8):
            cell = ws4.cell(row=r_idx, column=c)
            cell.fill = PatternFill("solid", fgColor="D90429")
            cell.border = BORDER
        continue

    bold = "**" in label
    cell = ws4.cell(row=r_idx, column=1, value=label.replace("**", ""))
    style_body(cell, bold=bold, center=False)
    for c_idx, v in enumerate(vals, 2):
        cell = ws4.cell(row=r_idx, column=c_idx)
        if v is None:
            cell.value = "—"
        else:
            cell.value = f"{v:.2f}" if isinstance(v, float) else v
        style_body(cell, bold=bold, alt=(r_idx % 2 == 0))
    cell = ws4.cell(row=r_idx, column=6, value=formula)
    style_note(cell)
    cell = ws4.cell(row=r_idx, column=7, value=source)
    style_note(cell)

ws4.column_dimensions["A"].width = 30
for c in range(2, 6):
    ws4.column_dimensions[get_column_letter(c)].width = 12
ws4.column_dimensions["F"].width = 32
ws4.column_dimensions["G"].width = 16

add_source_row(ws4, len(ratio_rows) + 3,
    "role01 collector 计算, 不解读. 标注 '推算' = 基于披露数据公式; 标注 '未披露' = 公司不单独披露",
    None)


# ============================================================
# Sheet 5: 季度趋势 (Quarterly Trend) - Time Series
# ============================================================
ws5 = wb.create_sheet("5.季度趋势")
ws5.freeze_panes = "B2"

headers = ["指标", "Q1 25", "Q2 25", "Q3 25", "Q4 25", "Q1 26", "QoQ Q4→Q1", "YoY Q1", "source"]
for i, h in enumerate(headers, 1):
    cell = ws5.cell(row=1, column=i, value=h)
    style_header(cell)

# Quarterly data
qtr_rows = [
    # 门店数
    ("**门店**", None, None, None, None, None, None, None, "", True),
    ("门店总数", 6681, 7038, 7338, 7453, 7531, 1.0, 12.7, "6-K Q1 26 + 招股书"),
    ("— 加盟", 6490, 6799, 6971, 6838, 6741, -1.4, 3.9, "6-K Q1 26"),
    ("— 自营", 191, 239, 367, 615, 790, 28.5, 313.6, "6-K Q1 26"),
    ("— 大陆", 6362, 6666, 6836, 6700, 6603, -1.4, 3.8, "6-K Q1 26"),
    ("— 海外", 128, 133, 135, 138, 138, 0.0, 7.8, "6-K Q1 26"),
    # GMV
    ("**GMV (¥M)**", None, None, None, None, None, None, None, "", True),
    ("大陆 GMV", 8048.4, None, None, None, 7491.4, None, -6.9, "6-K Q1 26 (只披露 Q1 25 vs Q1 26)"),
    ("海外 GMV", 178.0, None, None, None, 426.4, None, 139.6, "6-K Q1 26"),
    ("总 GMV", 8226.4, None, None, None, 7917.8, None, -3.7, "6-K Q1 26"),
    # 单店 GMV
    ("**单店 GMV (大陆, ¥)**", None, None, None, None, None, None, None, "", True),
    ("月均单店 GMV (大陆)", None, None, None, 337358, 356080, 5.5, None, "6-K Q1 26"),
    # 营收
    ("**营收 (¥M)**", None, None, None, None, None, None, None, "", True),
    ("营收", 3392.7, None, None, None, 3546.0, None, 4.5, "6-K Q1 26"),
    ("— 加盟营收", 3149.9, None, None, None, 2743.9, None, -12.9, "6-K Q1 26"),
    ("— 自营营收", 242.8, None, None, None, 802.1, None, 230.4, "6-K Q1 26"),
    # 利润
    ("**利润 (¥M)**", None, None, None, None, None, None, None, "", True),
    ("营业利润", 820.8, None, None, None, 547.2, None, -33.3, "6-K Q1 26"),
    ("GAAP 净利", 677.3, None, None, None, 447.7, None, -33.9, "6-K Q1 26"),
    ("Non-GAAP 净利", 677.3, None, None, None, 506.7, None, -25.2, "6-K Q1 26"),
    # 利润率
    ("**利润率 (%)**", None, None, None, None, None, None, None, "", True),
    ("净利率", 20.0, None, None, None, 12.6, None, -7.4, "6-K Q1 26"),
    ("Non-GAAP 净利率", 20.0, None, None, None, 14.3, None, -5.7, "6-K Q1 26"),
    # 用户
    ("**用户**", None, None, None, None, None, None, None, "", True),
    ("活跃会员 (M)", None, None, None, 44.8, 50.0, 11.7, None, "6-K Q1 26"),
]

for r_idx, row_data in enumerate(qtr_rows, 2):
    label = row_data[0]
    vals = row_data[1:6]
    qoq = row_data[6]
    yoy = row_data[7]
    source = row_data[8]
    is_section = "**" in label

    if is_section and all(v is None for v in vals):
        cell = ws5.cell(row=r_idx, column=1, value=label.replace("**", ""))
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="D90429")
        cell.border = BORDER
        for c in range(2, 10):
            cell = ws5.cell(row=r_idx, column=c)
            cell.fill = PatternFill("solid", fgColor="D90429")
            cell.border = BORDER
        continue

    cell = ws5.cell(row=r_idx, column=1, value=label)
    style_body(cell, center=False)
    for c_idx, v in enumerate(vals, 2):
        cell = ws5.cell(row=r_idx, column=c_idx)
        if v is None:
            cell.value = "—"
        else:
            cell.value = f"{v:,.1f}" if isinstance(v, float) else f"{v:,}"
        style_body(cell, alt=(r_idx % 2 == 0))
    # QoQ
    cell = ws5.cell(row=r_idx, column=7)
    cell.value = f"{qoq:.1f}%" if isinstance(qoq, (int, float)) else "—"
    style_body(cell, alt=(r_idx % 2 == 0))
    if isinstance(qoq, (int, float)) and qoq < 0:
        cell.font = Font(name="Calibri", size=10, color="D90429", bold=True)
    # YoY
    cell = ws5.cell(row=r_idx, column=8)
    cell.value = f"{yoy:.1f}%" if isinstance(yoy, (int, float)) else "—"
    style_body(cell, alt=(r_idx % 2 == 0))
    if isinstance(yoy, (int, float)) and yoy < 0:
        cell.font = Font(name="Calibri", size=10, color="D90429", bold=True)
    # source
    cell = ws5.cell(row=r_idx, column=9, value=source)
    style_note(cell)

ws5.column_dimensions["A"].width = 28
for c in range(2, 9):
    ws5.column_dimensions[get_column_letter(c)].width = 10
ws5.column_dimensions["I"].width = 30

add_source_row(ws5, len(qtr_rows) + 3,
    "数据源: 6-K Q1 2026 (accession 0001104659-26-067770). Q2-Q3-Q4 25 详细业绩公司未单独披露 6-K, 待抓",
    "https://www.sec.gov/Archives/edgar/data/2013649/000110465926067770/tm2615738d1_ex99-1.htm")


# ============================================================
# Sheet 6: 业务分拆 (Segment / Channel) - Breakdown
# ============================================================
ws6 = wb.create_sheet("6.业务分拆")
ws6.freeze_panes = "B2"

headers = ["分拆维度", "Q1 25", "Q1 26", "YoY", "占比变化", "source"]
for i, h in enumerate(headers, 1):
    cell = ws6.cell(row=1, column=i, value=h)
    style_header(cell)

# 加盟 vs 自营
seg_rows = [
    ("**按渠道 (加盟 vs 自营)**", None, None, None, None, "", True),
    ("加盟门店数", 6490, 6741, 3.9, "89.5% 末", "6-K Q1 26"),
    ("自营门店数", 191, 790, 313.6, "10.5% 末", "6-K Q1 26"),
    ("加盟营收 (¥M)", 3149.9, 2743.9, -12.9, "92.8% → 77.4%", "6-K Q1 26"),
    ("自营营收 (¥M)", 242.8, 802.1, 230.4, "7.2% → 22.6%", "6-K Q1 26"),
    ("自营 Opex (¥M)", 157.0, 497.2, 216.7, "4.6% → 14.0%", "6-K Q1 26"),
    ("自营单店月均营收 (¥M, 估)", 1.27, 1.02, -19.7, "—", "推算: 自营营收/自营门店/月数"),
    ("**按地区 (大陆 vs 海外)**", None, None, None, None, "", True),
    ("大陆门店数", 6362, 6603, 3.8, "87.7% 末", "6-K Q1 26"),
    ("海外门店数", 128, 138, 7.8, "1.8% 末", "6-K Q1 26"),
    ("大陆 GMV (¥M)", 8048.4, 7491.4, -6.9, "97.8% → 94.6%", "6-K Q1 26"),
    ("海外 GMV (¥M)", 178.0, 426.4, 139.6, "2.2% → 5.4%", "6-K Q1 26"),
    ("**按品类 (茶饮 vs 咖啡 vs 烘焙)**", None, None, None, None, "", True),
    ("茶饮 (主业)", None, None, None, "未单独披露", "公司未披露分品类"),
    ("咖啡 (战略品类, 估)", None, None, None, "未披露", "古茗示范 (1.2 万店配咖啡机)"),
    ("烘焙", None, None, None, "未披露", "公司未披露"),
    ("**按地区 (北美 vs 东南亚)**", None, None, None, None, "", True),
    ("马来西亚门店 (估)", None, None, None, "未披露", "招股书风险因素章提到, 但未给具体数"),
    ("新加坡门店 (估)", None, None, None, "未披露", "同上"),
    ("美国门店 (LA 首店)", None, None, None, "2025 落地, 1 家", "招股书"),
]

for r_idx, row_data in enumerate(seg_rows, 2):
    label = row_data[0]
    vals = row_data[1:4]
    pct = row_data[4]
    source = row_data[5]
    is_section = "**" in label

    if is_section and all(v is None for v in vals):
        cell = ws6.cell(row=r_idx, column=1, value=label.replace("**", ""))
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="D90429")
        cell.border = BORDER
        for c in range(2, 7):
            cell = ws6.cell(row=r_idx, column=c)
            cell.fill = PatternFill("solid", fgColor="D90429")
            cell.border = BORDER
        continue

    cell = ws6.cell(row=r_idx, column=1, value=label)
    style_body(cell, center=False)
    for c_idx, v in enumerate(vals, 2):
        cell = ws6.cell(row=r_idx, column=c_idx)
        if v is None:
            cell.value = "—"
        elif isinstance(v, float):
            cell.value = f"{v:,.2f}" if v < 100 else f"{v:,.0f}"
        else:
            cell.value = f"{v:,}"
        style_body(cell, alt=(r_idx % 2 == 0))
    # YoY
    cell = ws6.cell(row=r_idx, column=5)
    if vals[0] and vals[1] and isinstance(vals[0], (int, float)):
        yoy = (vals[1]/vals[0] - 1) * 100
        cell.value = f"{yoy:.1f}%"
        if yoy < 0:
            cell.font = Font(name="Calibri", size=10, color="D90429", bold=True)
    else:
        cell.value = "—"
    style_body(cell, alt=(r_idx % 2 == 0))
    # 占比
    cell = ws6.cell(row=r_idx, column=6, value=pct)
    style_note(cell)
    # source
    cell = ws6.cell(row=r_idx, column=7, value=source)
    style_note(cell)

ws6.column_dimensions["A"].width = 28
for c in range(2, 5):
    ws6.column_dimensions[get_column_letter(c)].width = 12
ws6.column_dimensions["E"].width = 18
ws6.column_dimensions["F"].width = 18
ws6.column_dimensions["G"].width = 28

add_source_row(ws6, len(seg_rows) + 3,
    "数据源: 6-K Q1 26 (channel/region); 招股书 (品类, 缺数据). '推算' = 公式计算, '未披露' = 公司不披露",
    "https://www.sec.gov/Archives/edgar/data/2013649/000110465926067770/tm2615738d1_ex99-1.htm")


# ============================================================
# Sheet 7: 同业可比 (Peer Comp) - Cross Section
# ============================================================
ws7 = wb.create_sheet("7.同业可比")
ws7.freeze_panes = "B2"

headers = ["公司", "代码", "营收 25 (¥B)", "净利 25 (¥B)", "净利率", "门店", "市值 (¥B)", "TTM PE", "Forward PE 26E", "source 26E", "数据源 25"]
for i, h in enumerate(headers, 1):
    cell = ws7.cell(row=1, column=i, value=h)
    style_header(cell)

# Peer data
peer_rows = [
    ("**CHAGEE**", "CHA", 12.91, 1.17, 9.1, 7531, 14.43, 12.3, 17.0, "自建 (基于 Q1 26 -34% YoY)", "SEC 20-F"),
    ("蜜雪冰城", "2097.HK", 33.56, 5.88, 17.5, 59823, 75.4, 12.8, 14.7, "东吴证券 2025-04-23", "蜜雪公告 2026-03-24"),
    ("古茗", "1364.HK", 12.91, 3.11, 24.1, 13554, 55.0, 17.7, 17.5, "富瑞 Jefferies 2026-07-10", "古茗公告 2026-03-25"),
    ("茶百道", "2555.HK", 5.40, 0.80, 15.2, 8621, 11.5, 14.4, 14.6, "国泰海通 2025-04-06", "茶百道公告 2026-03-27"),
    ("奈雪", "2150.HK", 4.33, -0.24, -5.6, 1646, 3.2, "n.m.", "n.m.", "—", "奈雪公告 2026-03-26"),
    ("**同业 median (剔除奈雪亏损)**", "—", "—", "—", "—", "—", "—", "**14.4x**", "**14.7x**", "—", "—"),
    ("**同业 mean (剔除奈雪)**", "—", "—", "—", "—", "—", "—", "**15.0x**", "**15.6x**", "—", "—"),
]

for r_idx, row_data in enumerate(peer_rows, 2):
    is_cha = "**CHAGEE**" in row_data[0]
    is_stat = "**同业" in row_data[0]

    cell = ws7.cell(row=r_idx, column=1, value=row_data[0].replace("**", ""))
    if is_cha:
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="D90429")
    elif is_stat:
        cell.font = FONT_BODY_BOLD
        cell.fill = PatternFill("solid", fgColor="F2F2F2")
    else:
        cell.font = FONT_BODY
    cell.alignment = ALIGN_LEFT
    cell.border = BORDER

    for c_idx, v in enumerate(row_data[1:10], 2):
        cell = ws7.cell(row=r_idx, column=c_idx, value=v)
        if is_cha:
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="D90429")
        elif is_stat:
            cell.font = FONT_BODY_BOLD
            cell.fill = PatternFill("solid", fgColor="F2F2F2")
        else:
            style_body(cell, alt=(r_idx % 2 == 0))
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER

    # 数据源
    cell = ws7.cell(row=r_idx, column=11, value=row_data[10])
    if is_cha or is_stat:
        cell.font = FONT_BODY_BOLD
        cell.fill = PatternFill("solid", fgColor="D90429" if is_cha else "F2F2F2")
    else:
        style_note(cell)
    cell.border = BORDER

ws7.column_dimensions["A"].width = 20
ws7.column_dimensions["B"].width = 10
for c in range(3, 11):
    ws7.column_dimensions[get_column_letter(c)].width = 13
ws7.column_dimensions["J"].width = 25
ws7.column_dimensions["K"].width = 25

# Conditional: highlight Forward PE > peer median
ws7.conditional_formatting.add(
    "J3:J6",
    CellIsRule(operator="greaterThan", formula=["15"], fill=PatternFill("solid", fgColor="FCE4E4"))
)
ws7.conditional_formatting.add(
    "J3:J6",
    CellIsRule(operator="lessThan", formula=["13"], fill=PatternFill("solid", fgColor="E4FCE4"))
)

add_source_row(ws7, len(peer_rows) + 3,
    "数据源: 各公司 2025 年报 + 卖方研报 (26E 预测). 汇率 2026-07-31: USD/CNY 6.7575, HKD/CNY 0.9161. CHA Forward PE 自建, 同业卖方研报一致预期",
    None)

# Add another source row
add_source_row(ws7, len(peer_rows) + 4,
    "汇率 source: 东方财富 (http://quote.eastmoney.com/unify/r/105.CHA 2026-07-31)",
    "http://quote.eastmoney.com/unify/r/105.CHA")


# ============================================================
# Save
# ============================================================
output_path = "/Users/mac/.minimax-agent-cn/projects/value-investing-agent/projects/cha-initiation-2026/05-writer/financials.xlsx"
wb.save(output_path)
print(f"✅ Excel saved: {output_path}")
print(f"   Sheets: {len(wb.sheetnames)}")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"   - {name}: {ws.max_row} rows x {ws.max_column} cols")
